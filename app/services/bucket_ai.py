from __future__ import annotations

import base64
import csv
import json
import logging
import mimetypes
import zipfile
from datetime import UTC, datetime
from io import BytesIO, StringIO
from typing import Any
from uuid import UUID

import boto3
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pypdf import PdfReader
from sqlalchemy import or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.config import get_settings
from app.models.bucket import (
    Bucket,
    BucketActivityLog,
    BucketAIActionItem,
    BucketAIMessage,
    BucketAIReview,
    BucketDocumentTemplate,
    BucketFile,
    BucketNote,
    BucketRequestedDocument,
    BucketShare,
    BucketUploadLink,
    BucketVendorAccess,
)
from app.models.user import User
from app.services.ai.bedrock_client import get_client, model_heavy, model_light
from app.services.ai.usage import json_safe_metadata, tracked_messages_create

log = logging.getLogger(__name__)

MAX_FILE_BYTES = 10 * 1024 * 1024
MAX_REVIEW_ATTACHMENTS = 8
MAX_PDF_PAGES = 100
MAX_TEXT_FILE_CHARS = 16000
MAX_PDF_TEXT_PAGES = 80
MAX_PDF_TEXT_CHARS = 50000
MAX_SPREADSHEET_SHEETS = 6
MAX_SPREADSHEET_ROWS = 80
MAX_SPREADSHEET_COLS = 30
MAX_SPREADSHEET_TEXT_CHARS = 24000
MAX_CELL_CHARS = 200
MAX_ZIP_ENTRIES = 30
MAX_ZIP_FILE_BYTES = 50 * 1024 * 1024
MAX_ZIP_ENTRY_BYTES = 40 * 1024 * 1024
MAX_ZIP_TOTAL_EXTRACTED_BYTES = 50 * 1024 * 1024
CHAT_WIDGET_TYPES = {
    "upload_files",
    "entity_structure",
    "deal_profile",
    "real_estate_schedule",
    "referral",
    "run_review",
    "bankability_result",
    "book_call",
}

REVIEW_SYSTEM = """You are a senior commercial lending underwriter reviewing a secure document bucket.

Return ONLY JSON in this shape. Do not wrap the JSON in markdown fences.
{
  "executive_summary": "short plain-English summary",
  "available_documents": [{"file_name": "...", "document_type": "...", "summary": "..."}],
  "missing_or_incomplete_items": [{"title": "...", "detail": "...", "priority": "high|medium|low"}],
  "discrepancies": [{"title": "...", "detail": "...", "files": ["..."]}],
  "underwriter_questions": [{"question": "...", "route": "admin|uploader|share", "reason": "..."}],
  "proof_of_funds_financial_collateral_gaps": [{"title": "...", "detail": "..."}],
  "recommended_next_document_requests": [{"title": "...", "instructions": "...", "route": "admin|uploader|share", "rationale": "..."}],
  "per_file_summaries": [{"file_id": "...", "file_name": "...", "summary": "...", "red_flags": ["..."]}],
  "document_evidence_map": {
    "files": [
      {
        "file_id": "...",
        "file_name": "...",
        "ai_classification": "tax_return|current_p_and_l|bank_statement|collateral_debt_evidence|real_estate_schedule|floorplan_mca_inventory|other|unreadable",
        "supports": ["..."],
        "baseline_categories_supported": ["..."],
        "confidence": "high|medium|low",
        "limitations": ["..."]
      }
    ],
    "baseline_coverage": [
      {"category": "last 2 years tax returns", "status": "satisfied|partial|missing|unclear", "evidence": ["..."], "gap": "..."}
    ]
  },
  "next_best_action": {
    "type": "upload_files|real_estate_schedule|entity_structure|deal_profile|book_call|none",
    "title": "...",
    "detail": "...",
    "reason": "..."
  },
  "screening_stage": "stage_1_bankability",
  "probability_status": "Good probability - book call|Promising but needs one clarification|Not enough evidence yet|Poor probability based on current file",
  "program_fit": ["real_estate_backed", "no_collateral_dscr", "full_doc_commercial", "high_cost_debt_refinance", "unknown"],
  "confidence": "high|medium|low",
  "key_metrics": {
    "revenue_trend": "...",
    "ytd_annualized_revenue": null,
    "annualized_adjusted_deposits": null,
    "tax_return_revenue_vs_bank_deposits": "...",
    "estimated_ebitda_or_cash_flow": null,
    "estimated_debt_burden": null,
    "estimated_dscr": null,
    "nsf_or_overdraft_flags": "...",
    "requested_amount_reasonableness": "..."
  },
  "strengths": ["..."],
  "risks": ["..."],
  "one_next_step": "...",
  "booking_recommended": false,
  "bankability_assessment": {
    "status": "Bankable|Incomplete - cannot determine|Not bankable based on current file",
    "product_fit": ["DSCR", "Full-doc commercial", "Dealer financing with real estate collateral"],
    "reason": "...",
    "conditions": ["..."]
  }
}

Be specific. Flag missing proof of funds, unclear financials, mismatched names/dates/amounts, missing collateral documents, unreadable files, stale documents, and any question an underwriter would ask before approval.
If ai_context.review_type is "dealer_gatekeeper_v1", act as a strict Stage 1 public lead gatekeeper for a car dealer financing file. Stage 1 is preliminary bankability, not final approval. Use the minimum Stage 1 package first: last 2 years business tax returns, current-year/YTD P&L, last 6 months main operating bank statements, requested amount, detailed use of funds with amount breakdown, stated current monthly debt payments, and estimated credit score/tier. Do not ask for a full underwriting package unless Stage 1 shows good probability. Treat real estate collateral as the next targeted clarification after cash-flow context indicates a possible path: property address, estimated value, current mortgage balance, and ownership/entity relationship if unclear. Floorplan/MCA/inventory statements are conditional only when documents, bank statements, or chat answers indicate those obligations exist.
For dealer_gatekeeper_v1, set screening_stage to "stage_1_bankability" and choose exactly one probability_status from: "Good probability - book call", "Promising but needs one clarification", "Not enough evidence yet", or "Poor probability based on current file". Set booking_recommended true only when the file has good lending probability and a human call should be offered now. Do not use CALL_NOW, MAYBE, or DO_NOT_CALL labels.
For dealer_gatekeeper_v1, calculate key_metrics when the documents allow it: revenue trend, YTD annualized revenue, annualized adjusted deposits, tax return revenue vs bank deposit consistency, estimated EBITDA/cash flow, estimated debt burden, estimated DSCR, NSF/overdraft/negative-balance flags, and requested amount reasonableness. Use null or a short unavailable explanation when the file does not support the metric.
For dealer_gatekeeper_v1 contexts, uploaded files may be random or miscategorized by the client. Classify documents from the readable content and file name first; do not rely only on requested_document_id. If a mortgage note, payoff statement, amortization schedule, loan statement, or debt schedule is uploaded, treat it as partial real-estate/collateral evidence even when the selected category is wrong. Do not say every baseline category is missing when any uploaded file supports a baseline category. Instead say exactly what the uploaded files prove and what is still missing, for example: "I see collateral debt evidence, but not tax returns, P&L, bank statements, property values, or entity relationship explanation." Ask for estimated values or missing addresses for collateral files when needed.
For dealer_gatekeeper_v1 incomplete reviews, structure the judgment like a senior banking underwriter: (1) what the uploaded files prove, (2) whether the file appears fundable, preliminarily fundable subject to confirmation, not fundable, or cannot be determined, (3) what still blocks a credit decision, and (4) the single next best clarification or baseline upload. Do not automatically jump to LLC/entity clarification unless the uploaded evidence or user's message makes entity/account relationships the immediate blocker. Do not use robotic "all categories missing" language when collateral/debt evidence exists.
For dealer_gatekeeper_v1 contexts with strong collateral, tax, wage, cash-flow, or asset evidence but missing confirmation documents, use "Incomplete - cannot determine" as the formal status but explicitly say whether the file appears "preliminarily fundable subject to confirmation" in the reason/summary. Separate missing confirmation documents from true not-bankable blockers.
For dealer_gatekeeper_v1 contexts where multiple LLCs, owner entities, real-estate entities, or dealership entities appear, ask for one written explanation that covers: primary operating LLC, main operating bank account, which LLCs own the real estate, how money transfers between entities, and whether dealership revenue supports property debt. This should be a conversational underwriting question, not a form or widget.
For dealer_gatekeeper_v1 client-facing summaries, organize the content as a short underwriting memo, not a dense paragraph. The JSON fields should support this order: status, what the files prove, what still blocks a decision, and one next step. Do not combine multiple client tasks into one next_best_action. If LLC/entity workflow is unclear, make the immediate next step only the first clarification: primary operating LLC and main operating bank account. The follow-up about related LLC money flow should come after the client answers.
If ai_context.review_type is "real_estate_dscr_v1", act as a strict real-estate investor / DSCR underwriter. This is not a car dealer review. Never ask about dealership name, floorplan, MCA, inventory, dealer gross receipts, or dealer LLC workflow. Screen purchase, refinance, and cash-out investor files using rent support, PITIA, DSCR, LTV, equity, cash to close, property condition, occupancy, lease/rent roll, purchase contract or payoff, taxes, insurance, HOA, entity/vesting, and estimated credit tier.
For real_estate_dscr_v1, set screening_stage to "stage_1_dscr_property_screen" and choose exactly one probability_status from: "Good probability - book call", "Promising but needs one clarification", "Not enough evidence yet", or "Poor probability based on current file". Set booking_recommended true only when the property/rent/equity evidence supports a good preliminary probability and a human call should be offered now.
For real_estate_dscr_v1, calculate key_metrics when documents allow it: DSCR, LTV, estimated property value, requested loan amount, equity, monthly rent, PITIA, NOI, cash to close, max supportable loan, reserve/cash gap, and credit-tier impact. Use null or a short unavailable explanation when evidence does not support a metric. Never invent numbers.
For real_estate_dscr_v1, classify uploaded files by readable content and filename. Treat leases, rent rolls, appraisal/value evidence, purchase contracts, payoff statements, mortgage statements, tax bills, insurance declarations, HOA statements, and settlement statements as useful partial evidence even when the selected category is wrong.
For real_estate_dscr_v1 client-facing summaries, use short sections and bullets: "What I see", "What it means", and "Next step". Ask only one next question or upload request at a time. If evidence is missing, say exactly which DSCR/LTV/PITIA/cash-to-close assumption is unsupported.
Keep the response compact enough to parse: executive_summary <= 1200 characters; available_documents <= 8 items; missing_or_incomplete_items <= 12 items; discrepancies <= 8 items; underwriter_questions <= 8 items; proof_of_funds_financial_collateral_gaps <= 8 items; recommended_next_document_requests <= 12 items; per_file_summaries <= 5 items; document_evidence_map.files <= 12 items; document_evidence_map.baseline_coverage <= 8 items. Keep each item detail/summary under 220 characters. Never list every file outside document_evidence_map. Per-file summaries are optional and should cover only the most important readable documents. Prioritize critical underwriting issues and one requested next action over exhaustive document recaps.
"""

CHAT_SYSTEM = """You are the Bucket AI assistant for a secure Qualified Commercial document room.

Return ONLY JSON in this shape:
{
  "answer": "helpful answer scoped to the user's permitted context",
  "proposed_context_patch": null,
  "proposed_action_items": [
    {"title": "...", "instructions": "...", "route": "admin|uploader|share|vendor", "rationale": "..."}
  ]
}

Rules:
- Never mention or infer files outside the provided context.
- External users cannot update saved bucket instructions directly.
- If an external user asks for something actionable, create proposed_action_items for super-admin approval.
- For admin users, proposed_context_patch may include deal_type, documentation_level, collateral_type, loan_purpose, underwriting_focus, or custom_instructions when the admin asks to update instructions.
- For admin users, when they ask you to create to-dos, missing-file requests, clarification requests, or follow-up actions, return those as proposed_action_items. Use route "uploader" for client upload tasks, "share" for one-time shared-reviewer tasks, "vendor" for authenticated vendor tasks, and "admin" for internal Qualified Commercial tasks.
- When suggesting document requests, prefer the provided document template names/categories when they fit. If none fit, create a clear custom task title and instructions.
- For dealer_gatekeeper_v1 contexts, Stage 1 asks for last 2 years business tax returns, YTD P&L, last 6 months main operating bank statements, requested amount, detailed use of funds with amount breakdown, stated current monthly debt payments, and estimated credit tier/score. Ask for one missing Stage 1 item at a time.
- When asking for use of funds, request a concrete breakdown: payoff amounts, working capital, inventory, taxes, repairs, acquisitions, cash-out reserves, or other planned uses.
- Do not ask for the full Stage 2 underwriting package unless the latest review says "Good probability - book call" or the user asks what comes after a promising Stage 1 screen.
- Treat real estate address, estimated value, and current mortgage balance as targeted follow-up clarifications after cash-flow context supports a possible real-estate-backed path.
- For dealer_gatekeeper_v1 contexts, use the latest evidence map if present. Explain what the uploaded files actually support before saying what is missing.
- When evidence is incomplete, choose only one next best action. Do not ask the user for every missing item at once.
- Write the answer like a human senior banking underwriter who understands used-car dealerships, real-estate-backed commercial lending, floorplan lines, MCA exposure, dealer cash flow, related LLCs, and operating account analysis.
- Do not sound like a generic chatbot. Avoid filler such as "I understand", "as an AI", "happy to help", and long tutorials.
- Be direct, calm, and practical: acknowledge the fact in front of you, state what it means for bankability, then give the next specific move.
- For dealer_gatekeeper_v1 contexts, use strict underwriting language: "I can preliminarily screen this", "this is incomplete", "this supports the file", "this creates a lender question", or "I cannot determine yet" when evidence is missing.
- Ask one high-value question at a time unless the user asks for a list.
- For dealer_gatekeeper_v1 chat answers, use short sections and bullets when summarizing files: "What I see", "What it means", and "Next step". Keep the next step to one client action.
- If related LLCs/accounts need clarification, ask only the first step first: primary operating LLC and main operating bank account. After the client answers, then ask how the related LLCs move money with the dealership.
- For real_estate_dscr_v1 contexts, answer like a real-estate investor / DSCR underwriter. Focus on rent, PITIA, DSCR, LTV, property value, payoff/purchase evidence, cash-to-close, entity/vesting, property condition, occupancy, and estimated credit tier.
- For real_estate_dscr_v1 contexts, do not mention dealership, floorplan, MCA, inventory, dealer gross receipts, or dealer operating accounts.
- For real_estate_dscr_v1 chat answers, use "What I see", "What it means", and "Next step" sections when summarizing uploaded evidence. Ask one real-estate funding question or upload request at a time.
- When the user asks to upload files or enter property collateral, answer directly in chat. Do not reference widgets, tools, sidebars, or navigation.
- Keep answers concise and operational, normally 2-5 short sentences.
"""


def _now() -> datetime:
    return datetime.now(UTC)


def _public_ai_context(bucket: Bucket) -> dict[str, Any]:
    context = bucket.ai_context or {}
    review_type = context.get("review_type")
    if review_type == "real_estate_dscr_v1":
        return {
            "review_type": "real_estate_dscr_v1",
            "deal_type": context.get("deal_type") or "real estate investor funding review",
            "documentation_level": context.get("documentation_level") or "preliminary DSCR / investor review",
            "collateral_type": context.get("collateral_type") or "income-producing real estate",
            "underwriter_persona": (
                "Speak like a senior real-estate investor and DSCR lending underwriter. "
                "Be practical, direct, and evidence-driven."
            ),
            "real_estate_knowledge": [
                "DSCR depends on supported rent and PITIA, not generic borrower claims.",
                "LTV depends on credible value evidence and current debt or requested loan amount.",
                "Purchase, refinance, and cash-out files need different evidence: contract, payoff, mortgage statement, taxes, insurance, HOA, rent support, and entity/vesting.",
            ],
            "baseline_document_policy": {
                "stage": "stage_1_dscr_property_screen",
                "allowed_document_categories": [
                    "lease, rent roll, or rent support",
                    "purchase contract for acquisitions",
                    "payoff or mortgage statement for refinance or cash-out",
                    "property tax, insurance, HOA, or PITIA support",
                    "entity or vesting documents when available",
                    "estimated value or purchase price",
                    "requested loan amount",
                    "monthly rent",
                    "transaction type",
                    "estimated credit tier",
                ],
                "do_not_request_dealer_documents": True,
                "stage_1_metrics": [
                    "DSCR",
                    "LTV",
                    "estimated property value",
                    "loan amount",
                    "equity",
                    "monthly rent",
                    "PITIA",
                    "NOI",
                    "cash to close",
                    "max supportable loan",
                    "reserve/cash gap",
                    "credit-tier impact",
                ],
            },
        }
    if review_type != "dealer_gatekeeper_v1":
        return {}
    return {
        "review_type": "dealer_gatekeeper_v1",
        "deal_type": context.get("deal_type") or "dealer financing with real estate collateral",
        "documentation_level": context.get("documentation_level") or "full doc",
        "collateral_type": context.get("collateral_type") or "real estate collateral and business assets",
        "underwriter_persona": (
            "Speak like a senior banking underwriter for used-car dealership financing. "
            "Be conversational but strict, practical, and evidence-driven."
        ),
        "dealer_knowledge": [
            "Used-car dealers often operate through multiple LLCs that share cash flow, inventory, real estate, and bank accounts.",
            "The main operating entity, main operating bank account, related entities, and account transfers must be clear before a confident screen.",
            "Floorplan debt, MCA debt, inventory turn, bank deposits, tax returns, P&L, and real estate equity all affect bankability.",
        ],
        "baseline_document_policy": {
            "stage": "stage_1_bankability",
            "allowed_document_categories": [
                "last 2 years business tax returns",
                "current year/YTD P&L",
                "last 6 months main operating bank statements",
                "requested amount",
                "detailed use of funds with amount breakdown",
                "stated current monthly debt payments",
                "estimated credit tier/score",
            ],
            "do_not_request_other_document_categories": True,
            "stage_1_metrics": [
                "revenue trend",
                "YTD annualized revenue",
                "annualized adjusted deposits",
                "tax return revenue vs bank deposits",
                "estimated EBITDA/cash flow",
                "estimated debt burden",
                "estimated DSCR",
                "NSF/overdraft/negative balance flags",
                "requested amount reasonableness",
            ],
            "conditional_follow_up_only_after_cash_flow_context": [
                "real estate full address",
                "real estate estimated property value",
                "real estate current mortgage balance",
                "primary operating LLC/entity",
                "main operating bank account",
                "related LLC/entity relationship",
                "floorplan/MCA/inventory statements when evidence suggests those obligations",
            ],
        },
        "answer_style": {
            "tone": "human banking underwriter, not generic chatbot",
            "format": "short sections and bullets for review summaries; 2-5 short sentences for ordinary chat",
            "client_pacing": "ask for one client action at a time; split LLC/account workflow into sequential clarifications",
            "avoid": ["generic AI disclaimers", "long tutorials", "open-ended document hunting"],
        },
    }


def _text_from_response(resp: Any) -> str:
    return "".join(block.text for block in resp.content if getattr(block, "type", None) == "text")


def _strip_code_fence(text: str) -> str:
    cleaned = (text or "").strip()
    if cleaned.startswith("```"):
        first_nl = cleaned.find("\n")
        if first_nl > 0:
            cleaned = cleaned[first_nl + 1 :]
        if cleaned.endswith("```"):
            cleaned = cleaned[:-3]
    return cleaned.strip()


def _without_optional_json_tail(text: str) -> str | None:
    cleaned = _strip_code_fence(text)
    if not cleaned.startswith("{"):
        return None
    optional_tail_keys = ("per_file_summaries", "available_documents")
    for key in optional_tail_keys:
        idx = cleaned.find(f'"{key}"')
        if idx <= 0:
            continue
        prefix = cleaned[:idx].rstrip()
        if prefix.endswith(":"):
            continue
        prefix = prefix.rstrip(", \n\r\t")
        if not prefix.endswith("{"):
            return prefix + "\n}"
    return None


def _json_or_fallback(text: str, fallback_key: str) -> dict[str, Any]:
    cleaned = _strip_code_fence(text)
    candidates = [cleaned]
    salvaged = _without_optional_json_tail(cleaned)
    if salvaged:
        candidates.append(salvaged)
    for candidate in candidates:
        try:
            parsed = json.loads(candidate)
            return parsed if isinstance(parsed, dict) else {fallback_key: text}
        except json.JSONDecodeError:
            continue
    return {fallback_key: text}


def _suggested_widget_type(parsed: dict[str, Any]) -> str | None:
    raw = parsed.get("suggested_widget")
    if isinstance(raw, dict):
        value = raw.get("type")
    else:
        value = raw
    if not isinstance(value, str):
        return None
    widget_type = value.strip()
    return widget_type if widget_type in CHAT_WIDGET_TYPES else None


def _ensure_review_result_shape(result: dict[str, Any]) -> dict[str, Any]:
    shaped = dict(result)
    shaped["screening_stage"] = str(shaped.get("screening_stage") or "stage_1_bankability")
    probability = str(shaped.get("probability_status") or "").strip()
    allowed_probability = {
        "Good probability - book call",
        "Promising but needs one clarification",
        "Not enough evidence yet",
        "Poor probability based on current file",
    }
    if probability not in allowed_probability:
        assessment = shaped.get("bankability_assessment") if isinstance(shaped.get("bankability_assessment"), dict) else {}
        combined = f"{assessment.get('status') or ''} {assessment.get('reason') or ''} {shaped.get('executive_summary') or ''}".lower()
        if any(term in combined for term in ("not bankable", "not fundable", "poor probability", "decline")):
            probability = "Poor probability based on current file"
        elif any(term in combined for term in ("good probability", "book call", "bankable", "fundable")) and not any(term in combined for term in ("cannot", "missing", "incomplete", "subject to")):
            probability = "Good probability - book call"
        elif any(term in combined for term in ("promising", "subject to", "one clarification", "preliminarily fundable")):
            probability = "Promising but needs one clarification"
        else:
            probability = "Not enough evidence yet"
    shaped["probability_status"] = probability
    shaped["booking_recommended"] = bool(shaped.get("booking_recommended") is True or probability == "Good probability - book call")
    if not isinstance(shaped.get("program_fit"), list):
        shaped["program_fit"] = [str(shaped.get("program_fit"))] if shaped.get("program_fit") else []
    confidence = str(shaped.get("confidence") or "").lower()
    shaped["confidence"] = confidence if confidence in {"high", "medium", "low"} else "low"
    if not isinstance(shaped.get("key_metrics"), dict):
        shaped["key_metrics"] = {}
    if not isinstance(shaped.get("strengths"), list):
        shaped["strengths"] = []
    if not isinstance(shaped.get("risks"), list):
        shaped["risks"] = []
    if not isinstance(shaped.get("one_next_step"), str):
        shaped["one_next_step"] = ""
    evidence_map = shaped.get("document_evidence_map")
    if not isinstance(evidence_map, dict):
        evidence_map = {"files": [], "baseline_coverage": []}
    else:
        evidence_map = dict(evidence_map)
        if not isinstance(evidence_map.get("files"), list):
            evidence_map["files"] = []
        if not isinstance(evidence_map.get("baseline_coverage"), list):
            evidence_map["baseline_coverage"] = []
    shaped["document_evidence_map"] = evidence_map
    next_best = shaped.get("next_best_action")
    if not isinstance(next_best, dict):
        recommended = shaped.get("recommended_next_document_requests")
        first = recommended[0] if isinstance(recommended, list) and recommended and isinstance(recommended[0], dict) else None
        detail = str(first.get("instructions") or first.get("detail") or "") if first else "Continue the underwriting conversation."
        next_best = {
            "type": "upload_files" if first else "none",
            "title": str(first.get("title") or "Next baseline item") if first else "No immediate tool needed",
            "detail": detail,
            "reason": str(first.get("rationale") or "Recommended by the review") if first else "No next best action was provided by the review.",
        }
    shaped["next_best_action"] = next_best
    return shaped


def _s3_client():
    settings = get_settings()
    kwargs: dict[str, Any] = {"region_name": settings.aws_region}
    if settings.aws_access_key_id and settings.aws_secret_access_key:
        kwargs["aws_access_key_id"] = settings.aws_access_key_id
        kwargs["aws_secret_access_key"] = settings.aws_secret_access_key
    return boto3.client("s3", **kwargs)


def _fetch_file(file: BucketFile) -> tuple[bytes, str] | None:
    settings = get_settings()
    if not settings.s3_bucket:
        return None
    try:
        obj = _s3_client().get_object(Bucket=settings.s3_bucket, Key=file.s3_key)
        return obj["Body"].read(), obj.get("ContentType") or file.content_type
    except Exception as exc:  # noqa: BLE001
        log.warning("bucket_ai: S3 fetch failed file=%s key=%s: %s", file.id, file.s3_key, exc)
        return None


def _media_type(content_type: str, file_name: str) -> str | None:
    lower = f"{content_type} {file_name}".lower()
    if "application/pdf" in lower or file_name.lower().endswith(".pdf"):
        return "application/pdf"
    if "image/jpeg" in lower or file_name.lower().endswith((".jpg", ".jpeg")):
        return "image/jpeg"
    if "image/png" in lower or file_name.lower().endswith(".png"):
        return "image/png"
    if "image/gif" in lower or file_name.lower().endswith(".gif"):
        return "image/gif"
    if "image/webp" in lower or file_name.lower().endswith(".webp"):
        return "image/webp"
    return None


def _content_block(media_type: str, raw: bytes) -> dict[str, Any]:
    encoded = base64.b64encode(raw).decode("ascii")
    if media_type == "application/pdf":
        return {"type": "document", "source": {"type": "base64", "media_type": media_type, "data": encoded}}
    return {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": encoded}}


def _pdf_review_metadata(raw: bytes) -> tuple[int | None, tuple[str, str] | None]:
    try:
        reader = PdfReader(BytesIO(raw), strict=False)
        if reader.is_encrypted:
            return (
                None,
                (
                    "password_protected",
                    "This PDF requires a password before AI can read it. Upload an unlocked copy or provide a readable replacement.",
                ),
            )
        page_count = len(reader.pages)
        if page_count > MAX_PDF_PAGES:
            return (
                page_count,
                (
                    "too_many_pdf_pages",
                    f"This PDF has {page_count} pages. Bedrock accepts PDFs up to {MAX_PDF_PAGES} pages, so this file was reviewed by metadata only.",
                ),
            )
        return page_count, None
    except Exception as exc:  # noqa: BLE001
        message = str(exc).lower()
        if "encrypt" in message or "password" in message:
            return (
                None,
                (
                    "password_protected",
                    "This PDF requires a password before AI can read it. Upload an unlocked copy or provide a readable replacement.",
                ),
            )
        return (
            None,
            (
                "pdf_parse_failed",
                "The system could not inspect this PDF safely before AI review, so it was reviewed by metadata only.",
            ),
        )


def _extract_pdf_text(file: BucketFile, raw: bytes, *, display_name: str | None = None) -> tuple[str | None, tuple[str, str] | None]:
    file_name = display_name or file.file_name
    try:
        reader = PdfReader(BytesIO(raw), strict=False)
        if reader.is_encrypted:
            return None, (
                "password_protected",
                "This PDF requires a password before AI can read it. Upload an unlocked copy or provide a readable replacement.",
            )
        page_count = len(reader.pages)
        pages: list[dict[str, Any]] = []
        chars_used = 0
        pages_examined = 0
        text_pages = 0
        truncated = False
        for page_offset in range(min(page_count, MAX_PDF_TEXT_PAGES)):
            page_index = page_offset + 1
            page = reader.pages[page_offset]
            pages_examined += 1
            try:
                page_text = (page.extract_text() or "").strip()
            except Exception as exc:  # noqa: BLE001
                log.warning("bucket_ai: PDF page text extract failed file=%s page=%s: %s", file.id, page_index, exc)
                continue
            if not page_text:
                continue
            text_pages += 1
            remaining = MAX_PDF_TEXT_CHARS - chars_used
            if remaining <= 0:
                truncated = True
                break
            if len(page_text) > remaining:
                page_text = page_text[:remaining]
                truncated = True
            chars_used += len(page_text)
            pages.append({"page": page_index, "text": page_text})
            if truncated:
                break
        if not pages:
            return None, (
                "pdf_text_unavailable",
                "This PDF could not be text-extracted. It may be scanned or image-only; upload a searchable PDF or split/readable copy.",
            )
        warnings = [
            warning
            for warning in (
                f"Only the first {MAX_PDF_TEXT_PAGES} pages were inspected." if page_count > MAX_PDF_TEXT_PAGES else None,
                f"Only the first {MAX_PDF_TEXT_CHARS} extracted characters were included." if truncated else None,
                "Text extraction may miss tables, stamps, signatures, or image-only pages.",
            )
            if warning
        ]
        payload = {
            "file_id": str(file.id),
            "file_name": file_name,
            "type": "large_pdf_text_extract",
            "page_count": page_count,
            "pages_examined": pages_examined,
            "pages_with_text": text_pages,
            "characters_included": chars_used,
            "warnings": warnings,
            "pages": pages,
        }
        return json.dumps(payload, default=str, ensure_ascii=False), None
    except Exception as exc:  # noqa: BLE001
        log.warning("bucket_ai: PDF text extraction failed file=%s: %s", file.id, exc)
        message = str(exc).lower()
        if "encrypt" in message or "password" in message:
            return None, (
                "password_protected",
                "This PDF requires a password before AI can read it. Upload an unlocked copy or provide a readable replacement.",
            )
        return None, (
            "pdf_text_extract_failed",
            "The system could not extract readable text from this PDF, so it was reviewed by metadata only.",
        )


def _is_xlsx_file(content_type: str, file_name: str) -> bool:
    lower = f"{content_type} {file_name}".lower()
    return (
        file_name.lower().endswith(".xlsx")
        or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in lower
    )


def _is_legacy_xls_file(file_name: str) -> bool:
    return file_name.lower().endswith(".xls")


def _is_csv_file(content_type: str, file_name: str) -> bool:
    lower = f"{content_type} {file_name}".lower()
    return file_name.lower().endswith(".csv") or "text/csv" in lower or "application/csv" in lower


def _compact_cell(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ").strip()
    if len(text) > MAX_CELL_CHARS:
        return text[: MAX_CELL_CHARS - 1] + "…"
    return text


def _trim_empty_tail(values: list[str]) -> list[str]:
    trimmed = list(values)
    while trimmed and not trimmed[-1]:
        trimmed.pop()
    return trimmed


def _decode_tabular_text(raw: bytes) -> tuple[str, str]:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return raw.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace"), "utf-8-replace"


def _extract_csv_text(file: BucketFile, raw: bytes, *, display_name: str | None = None) -> tuple[str | None, tuple[str, str] | None]:
    file_name = display_name or file.file_name
    try:
        decoded, encoding = _decode_tabular_text(raw)
        sample = decoded[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel
        rows: list[list[str]] = []
        truncated_rows = False
        truncated_cols = False
        for index, row in enumerate(csv.reader(StringIO(decoded), dialect)):
            if index >= MAX_SPREADSHEET_ROWS:
                truncated_rows = True
                break
            if len(row) > MAX_SPREADSHEET_COLS:
                truncated_cols = True
            compacted = [_compact_cell(value) for value in row[:MAX_SPREADSHEET_COLS]]
            rows.append(_trim_empty_tail(compacted))
        if not rows:
            return None, ("csv_parse_failed", "This CSV did not contain readable rows for AI review.")
        payload = {
            "file_id": str(file.id),
            "file_name": file_name,
            "type": "csv_table",
            "encoding": encoding,
            "rows_sampled": len(rows),
            "columns_sampled_limit": MAX_SPREADSHEET_COLS,
            "warnings": [
                warning
                for warning in (
                    f"Only the first {MAX_SPREADSHEET_ROWS} rows were included." if truncated_rows else None,
                    f"Only the first {MAX_SPREADSHEET_COLS} columns were included." if truncated_cols else None,
                )
                if warning
            ],
            "rows": rows,
        }
        return json.dumps(payload, default=str, ensure_ascii=False), None
    except Exception as exc:  # noqa: BLE001
        log.warning("bucket_ai: CSV parse failed file=%s: %s", file.id, exc)
        return None, ("csv_parse_failed", "The system could not parse this CSV safely, so it was reviewed by metadata only.")


def _extract_xlsx_text(file: BucketFile, raw: bytes, *, display_name: str | None = None) -> tuple[str | None, tuple[str, str] | None]:
    file_name = display_name or file.file_name
    values_wb = None
    formulas_wb = None
    try:
        values_wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
        formulas_wb = load_workbook(BytesIO(raw), read_only=True, data_only=False)
        visible_sheet_names = [
            sheet_name
            for sheet_name in formulas_wb.sheetnames
            if getattr(formulas_wb[sheet_name], "sheet_state", "visible") == "visible"
        ]
        workbook_warnings: list[str] = []
        if not visible_sheet_names:
            return None, ("spreadsheet_parse_failed", "This workbook does not contain visible sheets for AI review.")
        if len(visible_sheet_names) > MAX_SPREADSHEET_SHEETS:
            workbook_warnings.append(f"Only the first {MAX_SPREADSHEET_SHEETS} visible sheets were included.")

        sheets: list[dict[str, Any]] = []
        for sheet_name in visible_sheet_names[:MAX_SPREADSHEET_SHEETS]:
            formula_ws = formulas_wb[sheet_name]
            value_ws = values_wb[sheet_name]
            max_row = formula_ws.max_row or 0
            max_col = formula_ws.max_column or 0
            sheet_warnings: list[str] = []
            if max_row == 0 or max_col == 0:
                sheet_warnings.append("Sheet appears empty.")
                sheets.append(
                    {
                        "name": sheet_name,
                        "dimensions": {"rows": max_row, "columns": max_col},
                        "warnings": sheet_warnings,
                        "rows": [],
                        "formulas": [],
                    }
                )
                continue
            if max_row > MAX_SPREADSHEET_ROWS:
                sheet_warnings.append(f"Only the first {MAX_SPREADSHEET_ROWS} rows were included.")
            if max_col > MAX_SPREADSHEET_COLS:
                sheet_warnings.append(f"Only the first {MAX_SPREADSHEET_COLS} columns were included.")

            rows: list[list[str]] = []
            formulas: list[dict[str, str]] = []
            formula_iter = formula_ws.iter_rows(
                max_row=min(max_row, MAX_SPREADSHEET_ROWS),
                max_col=min(max_col, MAX_SPREADSHEET_COLS),
            )
            value_iter = value_ws.iter_rows(
                max_row=min(max_row, MAX_SPREADSHEET_ROWS),
                max_col=min(max_col, MAX_SPREADSHEET_COLS),
            )
            for row_index, (formula_row, value_row) in enumerate(zip(formula_iter, value_iter, strict=False), start=1):
                compacted: list[str] = []
                has_value = False
                for col_index, (formula_cell, value_cell) in enumerate(zip(formula_row, value_row, strict=False), start=1):
                    formula_value = formula_cell.value
                    cached_value = value_cell.value
                    display_value = cached_value if cached_value is not None else formula_value
                    compacted.append(_compact_cell(display_value))
                    if display_value is not None:
                        has_value = True
                    if isinstance(formula_value, str) and formula_value.startswith("=") and len(formulas) < 20:
                        formulas.append(
                            {
                                "cell": f"{get_column_letter(col_index)}{row_index}",
                                "formula": _compact_cell(formula_value),
                                "cached_value": _compact_cell(cached_value),
                            }
                        )
                trimmed = _trim_empty_tail(compacted)
                if has_value and trimmed:
                    rows.append(trimmed)
            if formulas and len(formulas) >= 20:
                sheet_warnings.append("Only the first 20 formulas were included.")
            sheets.append(
                {
                    "name": sheet_name,
                    "dimensions": {"rows": max_row, "columns": max_col},
                    "warnings": sheet_warnings,
                    "rows": rows,
                    "formulas": formulas,
                }
            )

        payload = {
            "file_id": str(file.id),
            "file_name": file_name,
            "type": "xlsx_workbook",
            "sheet_names": formulas_wb.sheetnames,
            "visible_sheets_included": [sheet["name"] for sheet in sheets],
            "warnings": workbook_warnings,
            "sheets": sheets,
        }
        return json.dumps(payload, default=str, ensure_ascii=False), None
    except Exception as exc:  # noqa: BLE001
        log.warning("bucket_ai: XLSX parse failed file=%s: %s", file.id, exc)
        return None, ("spreadsheet_parse_failed", "The system could not parse this workbook safely, so it was reviewed by metadata only.")
    finally:
        for workbook in (values_wb, formulas_wb):
            if workbook is not None:
                try:
                    workbook.close()
                except Exception:  # noqa: BLE001
                    pass


def _limit_structured_text(text: str, remaining_chars: int) -> tuple[str | None, bool]:
    if remaining_chars <= 0:
        return None, True
    if len(text) <= remaining_chars:
        return text, False
    return text[:remaining_chars], True


def _skip_file(file: BucketFile, reason: str, explanation: str) -> dict[str, str]:
    return {
        "file_id": str(file.id),
        "file_name": file.file_name,
        "reason": reason,
        "explanation": explanation,
    }


def _skip_named_file(file: BucketFile, file_name: str, reason: str, explanation: str) -> dict[str, str]:
    skipped = _skip_file(file, reason, explanation)
    skipped["file_name"] = file_name
    return skipped


def _is_zip_file(content_type: str, file_name: str) -> bool:
    lower = f"{content_type} {file_name}".lower()
    return file_name.lower().endswith(".zip") or "application/zip" in lower or "application/x-zip-compressed" in lower


def _safe_zip_member_name(name: str) -> str | None:
    cleaned = name.replace("\\", "/").strip()
    if not cleaned or cleaned.startswith("/") or cleaned.endswith("/"):
        return None
    parts = [part for part in cleaned.split("/") if part]
    if not parts or any(part in {".", ".."} for part in parts):
        return None
    return "/".join(parts)


def _append_review_file_content(
    *,
    content: list[dict[str, Any]],
    skipped: list[dict[str, str]],
    blocked_files: list[dict[str, str]],
    file: BucketFile,
    raw: bytes,
    content_type: str,
    display_name: str | None = None,
    attached_pdf_pages: int,
    spreadsheet_text_chars: int,
) -> tuple[bool, int, int]:
    file_name = display_name or file.file_name
    media = _media_type(content_type, file_name)
    if media == "application/pdf" and len(raw) > MAX_FILE_BYTES:
        extracted, parse_skip = _extract_pdf_text(file, raw, display_name=file_name)
        if parse_skip:
            reason, explanation = parse_skip
            skipped_file = _skip_named_file(file, file_name, reason, explanation)
            if reason == "password_protected":
                blocked_files.append(skipped_file)
            skipped.append(skipped_file)
            return False, attached_pdf_pages, spreadsheet_text_chars
        content.append({"type": "text", "text": f"Large PDF file {file.id}: {file_name}\nExtracted searchable text for underwriting review because the file is too large for direct model attachment:\n\n{extracted}"})
        skipped.append(
            _skip_named_file(
                file,
                file_name,
                "large_pdf_text_extract",
                "The PDF was too large to attach directly, so searchable text was extracted and reviewed. Tables/images/signatures may still need a readable split copy if the AI flags gaps.",
            )
        )
        return True, attached_pdf_pages, spreadsheet_text_chars
    if len(raw) > MAX_FILE_BYTES:
        skipped.append(_skip_named_file(file, file_name, "too_large", "The file is larger than the current AI review limit and was reviewed by metadata only."))
        return False, attached_pdf_pages, spreadsheet_text_chars
    if _is_xlsx_file(content_type, file_name):
        extracted, parse_skip = _extract_xlsx_text(file, raw, display_name=file_name)
        if parse_skip:
            reason, explanation = parse_skip
            skipped.append(_skip_named_file(file, file_name, reason, explanation))
            return False, attached_pdf_pages, spreadsheet_text_chars
        limited, truncated = _limit_structured_text(extracted or "", MAX_SPREADSHEET_TEXT_CHARS - spreadsheet_text_chars)
        if limited is None:
            skipped.append(_skip_named_file(file, file_name, "spreadsheet_too_large", "The AI review reached the spreadsheet text budget, so this workbook was reviewed by metadata only."))
            return False, attached_pdf_pages, spreadsheet_text_chars
        spreadsheet_text_chars += len(limited)
        content.append({"type": "text", "text": f"Spreadsheet file {file.id}: {file_name}\nExtracted workbook/table data for underwriting review:\n\n{limited}"})
        if truncated:
            skipped.append(_skip_named_file(file, file_name, "spreadsheet_too_large", "Only the first part of this workbook could be included before the spreadsheet text budget was reached."))
        return True, attached_pdf_pages, spreadsheet_text_chars
    if _is_legacy_xls_file(file_name):
        skipped.append(_skip_named_file(file, file_name, "legacy_spreadsheet_unsupported", "Legacy .xls files are not supported by the current AI spreadsheet parser. Upload an .xlsx or CSV export."))
        return False, attached_pdf_pages, spreadsheet_text_chars
    if _is_csv_file(content_type, file_name):
        extracted, parse_skip = _extract_csv_text(file, raw, display_name=file_name)
        if parse_skip:
            reason, explanation = parse_skip
            skipped.append(_skip_named_file(file, file_name, reason, explanation))
            return False, attached_pdf_pages, spreadsheet_text_chars
        limited, truncated = _limit_structured_text(extracted or "", MAX_SPREADSHEET_TEXT_CHARS - spreadsheet_text_chars)
        if limited is None:
            skipped.append(_skip_named_file(file, file_name, "spreadsheet_too_large", "The AI review reached the spreadsheet text budget, so this CSV was reviewed by metadata only."))
            return False, attached_pdf_pages, spreadsheet_text_chars
        spreadsheet_text_chars += len(limited)
        content.append({"type": "text", "text": f"CSV file {file.id}: {file_name}\nExtracted table data for underwriting review:\n\n{limited}"})
        if truncated:
            skipped.append(_skip_named_file(file, file_name, "spreadsheet_too_large", "Only the first part of this CSV could be included before the spreadsheet text budget was reached."))
        return True, attached_pdf_pages, spreadsheet_text_chars
    if media:
        if media == "application/pdf":
            page_count, pdf_skip = _pdf_review_metadata(raw)
            if pdf_skip:
                reason, explanation = pdf_skip
                if reason in {"too_many_pdf_pages", "pdf_parse_failed"}:
                    extracted, extract_skip = _extract_pdf_text(file, raw, display_name=file_name)
                    if extracted:
                        content.append({"type": "text", "text": f"PDF file {file.id}: {file_name}\nExtracted searchable text for underwriting review because direct PDF attachment was not available ({reason}):\n\n{extracted}"})
                        skipped.append(
                            _skip_named_file(
                                file,
                                file_name,
                                "pdf_text_extract_used",
                                "The PDF could not be attached directly, so searchable text was extracted and reviewed. Tables/images/signatures may still need a readable split copy if the AI flags gaps.",
                            )
                        )
                        return True, attached_pdf_pages, spreadsheet_text_chars
                    if extract_skip:
                        reason, explanation = extract_skip
                skipped_file = _skip_named_file(file, file_name, reason, explanation)
                if reason == "password_protected":
                    blocked_files.append(skipped_file)
                skipped.append(skipped_file)
                return False, attached_pdf_pages, spreadsheet_text_chars
            if page_count is not None and attached_pdf_pages + page_count > MAX_PDF_PAGES:
                extracted, extract_skip = _extract_pdf_text(file, raw, display_name=file_name)
                if extracted:
                    content.append({"type": "text", "text": f"PDF file {file.id}: {file_name}\nExtracted searchable text for underwriting review because the total direct PDF page budget was reached:\n\n{extracted}"})
                    skipped.append(
                        _skip_named_file(
                            file,
                            file_name,
                            "pdf_page_budget_text_extract",
                            f"Direct PDF attachment was skipped because Bedrock accepts up to {MAX_PDF_PAGES} total PDF pages per review. Searchable text was extracted and reviewed instead.",
                        )
                    )
                    return True, attached_pdf_pages, spreadsheet_text_chars
                if extract_skip:
                    reason, explanation = extract_skip
                    skipped.append(_skip_named_file(file, file_name, reason, explanation))
                else:
                    skipped.append(_skip_named_file(file, file_name, "pdf_page_budget_exceeded", f"Bedrock accepts up to {MAX_PDF_PAGES} total PDF pages per review. This file would bring the review to {attached_pdf_pages + page_count} pages, so it was reviewed by metadata only."))
                return False, attached_pdf_pages, spreadsheet_text_chars
            attached_pdf_pages += page_count or 0
            content.append({"type": "text", "text": f"PDF file {file.id}: {file_name}"})
        else:
            content.append({"type": "text", "text": f"Image file {file.id}: {file_name} ({media}) attached for visual underwriting review."})
        content.append(_content_block(media, raw))
        return True, attached_pdf_pages, spreadsheet_text_chars
    lower = f"{content_type} {file_name}".lower()
    if "text/" in lower or file_name.lower().endswith((".txt", ".md", ".log")):
        snippet = raw[:MAX_TEXT_FILE_CHARS].decode("utf-8", errors="replace")
        content.append({"type": "text", "text": f"File {file.id}: {file_name}\n\n{snippet}"})
        return True, attached_pdf_pages, spreadsheet_text_chars
    skipped.append(_skip_named_file(file, file_name, "unsupported_content_type", "This file type is not directly attached to the AI model yet and was reviewed by metadata only."))
    return False, attached_pdf_pages, spreadsheet_text_chars


def _append_zip_file_content(
    *,
    content: list[dict[str, Any]],
    skipped: list[dict[str, str]],
    blocked_files: list[dict[str, str]],
    file: BucketFile,
    raw: bytes,
    attached: int,
    attached_pdf_pages: int,
    spreadsheet_text_chars: int,
) -> tuple[int, int, int]:
    try:
        with zipfile.ZipFile(BytesIO(raw)) as archive:
            members = archive.infolist()
            if len(members) > MAX_ZIP_ENTRIES:
                skipped.append(_skip_file(file, "zip_entry_limit", f"Only the first {MAX_ZIP_ENTRIES} ZIP entries can be inspected in one AI review."))
            total_extracted = 0
            inspected = 0
            content.append({"type": "text", "text": f"ZIP file {file.id}: {file.file_name} accepted. The AI will review supported files extracted from this archive."})
            for member in members:
                if inspected >= MAX_ZIP_ENTRIES or attached >= MAX_REVIEW_ATTACHMENTS:
                    break
                inspected += 1
                member_name = _safe_zip_member_name(member.filename)
                display_name = f"{file.file_name}::{member_name or member.filename}"
                if member_name is None:
                    skipped.append(_skip_named_file(file, display_name, "zip_unsafe_path", "This ZIP entry used an unsafe or empty path and was skipped."))
                    continue
                if member_name.lower().endswith(".zip"):
                    skipped.append(_skip_named_file(file, display_name, "nested_zip_unsupported", "Nested ZIP files are not extracted for AI review."))
                    continue
                if member.flag_bits & 0x1:
                    skipped.append(_skip_named_file(file, display_name, "zip_entry_encrypted", "This ZIP entry is encrypted and cannot be read by AI review."))
                    continue
                if member.file_size > MAX_ZIP_ENTRY_BYTES:
                    skipped.append(_skip_named_file(file, display_name, "zip_entry_too_large", "This ZIP entry is larger than the current AI review limit and was skipped."))
                    continue
                total_extracted += member.file_size
                if total_extracted > MAX_ZIP_TOTAL_EXTRACTED_BYTES:
                    skipped.append(_skip_named_file(file, display_name, "zip_total_too_large", "The ZIP extraction budget was reached, so remaining entries were skipped."))
                    break
                try:
                    member_raw = archive.read(member)
                except RuntimeError:
                    skipped.append(_skip_named_file(file, display_name, "zip_entry_encrypted", "This ZIP entry is encrypted and cannot be read by AI review."))
                    continue
                member_content_type = mimetypes.guess_type(member_name)[0] or "application/octet-stream"
                added, attached_pdf_pages, spreadsheet_text_chars = _append_review_file_content(
                    content=content,
                    skipped=skipped,
                    blocked_files=blocked_files,
                    file=file,
                    raw=member_raw,
                    content_type=member_content_type,
                    display_name=display_name,
                    attached_pdf_pages=attached_pdf_pages,
                    spreadsheet_text_chars=spreadsheet_text_chars,
                )
                if added:
                    attached += 1
    except zipfile.BadZipFile:
        skipped.append(_skip_file(file, "zip_parse_failed", "The uploaded ZIP could not be opened safely and was reviewed by metadata only."))
    return attached, attached_pdf_pages, spreadsheet_text_chars


async def log_bucket_ai_activity(
    db: AsyncSession,
    bucket_id: UUID,
    action: str,
    *,
    user: User | None = None,
    actor_name: str | None = None,
    actor_email: str | None = None,
    actor_role: str | None = None,
    target_type: str | None = None,
    target_id: str | None = None,
    detail: str | None = None,
) -> None:
    db.add(
        BucketActivityLog(
            bucket_id=bucket_id,
            actor_user_id=user.id if user else None,
            actor_name=user.name if user else actor_name,
            actor_email=user.email if user else actor_email,
            actor_role=user.role if user else actor_role,
            action=action,
            target_type=target_type,
            target_id=target_id,
            detail=detail,
            created_at=_now(),
        )
    )


async def run_bucket_ai_review(db: AsyncSession, review_id: UUID) -> BucketAIReview | None:
    review = (
        await db.execute(
            select(BucketAIReview)
            .where(BucketAIReview.id == review_id)
            .options(
                selectinload(BucketAIReview.bucket).selectinload(Bucket.requested_documents),
                selectinload(BucketAIReview.bucket).selectinload(Bucket.files),
            )
        )
    ).scalar_one_or_none()
    if review is None or review.status not in {"queued", "failed"}:
        return review

    bucket = review.bucket
    review.status = "running"
    review.started_at = _now()
    review.error = None
    await log_bucket_ai_activity(db, bucket.id, "ai_review_started", target_type="ai_review", target_id=str(review.id), detail=bucket.name)
    await db.flush()

    files = [file for file in bucket.files if file.status == "uploaded" and file.deleted_at is None]
    extracted_zip_parent_ids = {
        file.parent_zip_file_id
        for file in files
        if getattr(file, "parent_zip_file_id", None) is not None and file.deleted_at is None
    }
    review.file_ids = [str(file.id) for file in files]

    requested = [
        {
            "id": str(doc.id),
            "name": doc.name,
            "category": doc.category,
            "description": doc.description,
            "required": doc.required,
            "status": doc.status,
        }
        for doc in bucket.requested_documents
    ]
    metadata = [
        {
            "id": str(file.id),
            "file_name": file.file_name,
            "content_type": file.content_type,
            "size_bytes": file.size_bytes,
            "requested_document_id": str(file.requested_document_id) if file.requested_document_id else None,
            "uploaded_by": file.uploaded_by_name,
            "parent_zip_file_id": str(file.parent_zip_file_id) if getattr(file, "parent_zip_file_id", None) else None,
            "zip_entry_path": getattr(file, "zip_entry_path", None),
            "extraction_status": getattr(file, "extraction_status", None),
            "extraction_reason": getattr(file, "extraction_reason", None),
        }
        for file in files
    ]
    content: list[dict[str, Any]] = [
        {
            "type": "text",
            "text": json.dumps(
                {
                    "bucket": {
                        "name": bucket.name,
                        "client_name": bucket.client_name,
                        "purpose": bucket.purpose,
                        "bucket_type": bucket.bucket_type,
                        "description": bucket.description,
                    },
                    "ai_context": review.context_snapshot or bucket.ai_context or {},
                    "requested_documents": requested,
                    "uploaded_files": metadata,
                    "instruction": "Review the attached/readable files and the metadata. Identify what is available, missing, discrepant, unclear, or likely to be questioned by an underwriter.",
                },
                default=str,
            ),
        }
    ]

    attached = 0
    attached_pdf_pages = 0
    spreadsheet_text_chars = 0
    skipped: list[dict[str, str]] = []
    blocked_files: list[dict[str, str]] = []
    for file in files:
        if attached >= MAX_REVIEW_ATTACHMENTS:
            skipped.append(_skip_file(file, "attachment_limit", "The AI review reached the attachment limit, so this file was reviewed by metadata only."))
            continue
        fetched = _fetch_file(file)
        if fetched is None:
            skipped.append(_skip_file(file, "fetch_failed", "The system could not retrieve this file from storage for AI review."))
            continue
        raw, content_type = fetched
        if _is_zip_file(content_type, file.file_name):
            if file.id in extracted_zip_parent_ids:
                skipped.append(_skip_file(file, "zip_parent_archive", "ZIP contents were extracted into bucket files, so the archive itself was not re-read by AI review."))
                continue
            if len(raw) > MAX_ZIP_FILE_BYTES:
                skipped.append(_skip_file(file, "zip_too_large", "The ZIP file is larger than the current AI extraction limit and was reviewed by metadata only."))
                continue
            attached, attached_pdf_pages, spreadsheet_text_chars = _append_zip_file_content(
                content=content,
                skipped=skipped,
                blocked_files=blocked_files,
                file=file,
                raw=raw,
                attached=attached,
                attached_pdf_pages=attached_pdf_pages,
                spreadsheet_text_chars=spreadsheet_text_chars,
            )
            continue
        added, attached_pdf_pages, spreadsheet_text_chars = _append_review_file_content(
            content=content,
            skipped=skipped,
            blocked_files=blocked_files,
            file=file,
            raw=raw,
            content_type=content_type,
            attached_pdf_pages=attached_pdf_pages,
            spreadsheet_text_chars=spreadsheet_text_chars,
        )
        if added:
            attached += 1

    if skipped:
        content.append({"type": "text", "text": "Files not attached to model: " + json.dumps(skipped)})
    if blocked_files:
        content.append({"type": "text", "text": "Files requiring action before AI can read them: " + json.dumps(blocked_files)})

    try:
        model = model_heavy()
        resp = await tracked_messages_create(
            db,
            feature="document_scan",
            client=get_client(),
            model=model,
            metadata={"bucket_id": str(bucket.id), "bucket_ai_review_id": str(review.id)},
            max_tokens=6000,
            system=REVIEW_SYSTEM,
            messages=[{"role": "user", "content": content}],
        )
        review.provider = "bedrock"
        review.model = getattr(resp, "model", None) or model
        result = _ensure_review_result_shape(_json_or_fallback(_text_from_response(resp), "executive_summary"))
        if blocked_files:
            result["blocked_files"] = blocked_files
        if skipped:
            result["skipped_files"] = skipped
        review.result = result
        review.status = "completed"
        review.completed_at = _now()
        await _create_review_recommendation_actions(db, bucket=bucket, review=review, result=result)
        await log_bucket_ai_activity(db, bucket.id, "ai_review_completed", target_type="ai_review", target_id=str(review.id), detail=bucket.name)
    except Exception as exc:  # noqa: BLE001
        log.exception("bucket_ai: review failed review=%s", review.id)
        review.status = "failed"
        review.error = str(exc)[:2000]
        review.completed_at = _now()
        await log_bucket_ai_activity(db, bucket.id, "ai_review_failed", target_type="ai_review", target_id=str(review.id), detail=review.error)
    await db.flush()
    return review


async def drain_bucket_ai_reviews(db: AsyncSession, *, limit: int = 3) -> int:
    rows = (
        await db.execute(
            select(BucketAIReview.id)
            .where(BucketAIReview.status == "queued")
            .order_by(BucketAIReview.created_at.asc())
            .limit(limit)
        )
    ).scalars().all()
    for review_id in rows:
        await run_bucket_ai_review(db, review_id)
        await db.commit()
    return len(rows)


async def latest_review(db: AsyncSession, bucket_id: UUID) -> BucketAIReview | None:
    completed = (
        await db.execute(
            select(BucketAIReview)
            .where(BucketAIReview.bucket_id == bucket_id, BucketAIReview.status == "completed")
            .order_by(BucketAIReview.completed_at.desc().nulls_last(), BucketAIReview.created_at.desc())
            .limit(1)
        )
    ).scalar_one_or_none()
    if completed is not None:
        return completed
    return (
        await db.execute(
            select(BucketAIReview)
            .where(BucketAIReview.bucket_id == bucket_id, BucketAIReview.result.is_not(None))
            .order_by(BucketAIReview.completed_at.desc().nulls_last(), BucketAIReview.created_at.desc())
            .limit(1)
        )
    ).scalar_one_or_none()


def share_visible_summary(review: BucketAIReview | None, share: BucketShare) -> dict[str, Any] | None:
    if review is None or not isinstance(review.result, dict):
        return None
    visible_names = {file.file_name for file in share.files if file.status == "uploaded" and file.deleted_at is None}
    per_file = [
        item for item in review.result.get("per_file_summaries", []) or []
        if isinstance(item, dict) and item.get("file_name") in visible_names
    ]
    return {
        "summary": f"{len(visible_names)} shared file{'' if len(visible_names) == 1 else 's'} available for review.",
        "per_file_summaries": per_file,
        "missing_or_incomplete_items": _visible_review_items(review.result.get("missing_or_incomplete_items") or [], visible_names),
        "discrepancies": _visible_review_items(review.result.get("discrepancies") or [], visible_names),
        "blocked_files": _visible_review_items(review.result.get("blocked_files") or [], visible_names),
    }


def vendor_visible_summary(review: BucketAIReview | None, access: BucketVendorAccess) -> dict[str, Any] | None:
    if review is None or not isinstance(review.result, dict):
        return None
    visible_files = (
        access.bucket.files if access.file_scope == "all_active" else access.files
    )
    visible_names = {
        file.file_name for file in visible_files
        if file.status == "uploaded" and file.deleted_at is None
    }
    per_file = [
        item for item in review.result.get("per_file_summaries", []) or []
        if isinstance(item, dict) and item.get("file_name") in visible_names
    ]
    return {
        "summary": f"{len(visible_names)} vendor-visible file{'' if len(visible_names) == 1 else 's'} available for review.",
        "per_file_summaries": per_file,
        "missing_or_incomplete_items": _visible_review_items(review.result.get("missing_or_incomplete_items") or [], visible_names),
        "discrepancies": _visible_review_items(review.result.get("discrepancies") or [], visible_names),
        "blocked_files": _visible_review_items(review.result.get("blocked_files") or [], visible_names),
    }


def _visible_review_items(items: Any, visible_names: set[str]) -> list[dict[str, Any]]:
    if not isinstance(items, list):
        return []
    visible: list[dict[str, Any]] = []
    for item in items:
        if not isinstance(item, dict):
            continue
        raw_files = item.get("files")
        named_files = {str(value) for value in raw_files} if isinstance(raw_files, list) else set()
        file_name = str(item.get("file_name") or "")
        if named_files and not named_files.intersection(visible_names):
            continue
        if file_name and file_name not in visible_names:
            continue
        visible.append(item)
    return visible


def upload_link_visible_summary(review: BucketAIReview | None, bucket: Bucket) -> dict[str, Any] | None:
    active_files = [file for file in bucket.files if file.status == "uploaded" and file.deleted_at is None]
    active_names = {file.file_name for file in active_files}
    active_ids = {str(file.id) for file in active_files}
    if review is None or not isinstance(review.result, dict):
        missing_docs = [
            {
                "title": doc.name,
                "detail": doc.description or "This requested document has not been marked received yet.",
                "priority": "high" if doc.required else "medium",
                "category": doc.category,
            }
            for doc in bucket.requested_documents
            if doc.status != "uploaded"
        ]
        return {
            "summary": "Qualified Commercial has prepared this file room. AI analysis uses the super-admin bucket inputs, uploaded files, and any completed underwriting review.",
            "review_status": review.status if review else "not_started",
            "review_error": review.error if review else None,
            "ai_context": bucket.ai_context or {},
            "available_documents": [
                {
                    "file_id": str(file.id),
                    "file_name": file.file_name,
                    "document_type": file.content_type,
                    "summary": f"Uploaded by {file.uploaded_by_name or file.uploaded_by_email or 'Qualified Commercial'}.",
                }
                for file in active_files
            ],
            "missing_or_incomplete_items": missing_docs,
            "discrepancies": [],
            "underwriter_questions": [],
            "proof_of_funds_financial_collateral_gaps": [],
            "per_file_summaries": [],
            "blocked_files": [],
            "skipped_files": [],
        }
    result = review.result
    per_file = [
        item for item in result.get("per_file_summaries", []) or []
        if isinstance(item, dict)
        and (str(item.get("file_id") or "") in active_ids or str(item.get("file_name") or "") in active_names)
    ]
    available = [
        item for item in result.get("available_documents", []) or []
        if isinstance(item, dict) and (not item.get("file_name") or str(item.get("file_name")) in active_names)
    ]
    return {
        "summary": result.get("executive_summary") or result.get("summary") or f"{len(active_files)} uploaded file{'' if len(active_files) == 1 else 's'} available.",
        "ai_context": bucket.ai_context or {},
        "review_completed_at": review.completed_at.isoformat() if review.completed_at else None,
        "available_documents": available,
        "missing_or_incomplete_items": result.get("missing_or_incomplete_items") or [],
        "discrepancies": result.get("discrepancies") or [],
        "underwriter_questions": result.get("underwriter_questions") or [],
        "proof_of_funds_financial_collateral_gaps": result.get("proof_of_funds_financial_collateral_gaps") or [],
        "per_file_summaries": per_file,
        "blocked_files": result.get("blocked_files") or [],
        "skipped_files": result.get("skipped_files") or [],
    }


async def visible_action_items(
    db: AsyncSession,
    bucket_id: UUID,
    *,
    route: str | None = None,
    upload_link_id: UUID | None = None,
    share_id: UUID | None = None,
    vendor_access_id: UUID | None = None,
    approved_only: bool = False,
) -> list[BucketAIActionItem]:
    stmt = select(BucketAIActionItem).where(BucketAIActionItem.bucket_id == bucket_id)
    if approved_only:
        stmt = stmt.where(BucketAIActionItem.status.in_(("approved", "completed")))
    if route:
        stmt = stmt.where(BucketAIActionItem.route == route)
    if upload_link_id:
        stmt = stmt.where(
            or_(
                BucketAIActionItem.upload_link_id == upload_link_id,
                BucketAIActionItem.upload_link_id.is_(None),
            )
            if route == "uploader"
            else BucketAIActionItem.upload_link_id == upload_link_id
        )
    if share_id:
        stmt = stmt.where(BucketAIActionItem.share_id == share_id)
    if vendor_access_id:
        stmt = stmt.where(BucketAIActionItem.vendor_access_id == vendor_access_id)
    return (await db.execute(stmt.order_by(BucketAIActionItem.created_at.desc()))).scalars().all()


async def create_chat_reply(
    db: AsyncSession,
    *,
    bucket: Bucket,
    audience: str,
    message: str,
    actor_name: str,
    user: User | None = None,
    upload_link: BucketUploadLink | None = None,
    share: BucketShare | None = None,
    vendor_access: BucketVendorAccess | None = None,
) -> tuple[list[BucketAIMessage], list[BucketAIActionItem], str | None]:
    user_row = BucketAIMessage(
        bucket_id=bucket.id,
        upload_link_id=upload_link.id if upload_link else None,
        share_id=share.id if share else None,
        vendor_access_id=vendor_access.id if vendor_access else None,
        user_id=user.id if user else None,
        audience=audience,
        role="user",
        author_name=actor_name,
        content=message,
    )
    db.add(user_row)
    await db.flush()

    context = await _chat_context(db, bucket=bucket, audience=audience, upload_link=upload_link, share=share, vendor_access=vendor_access)
    model = model_light()
    try:
        resp = await tracked_messages_create(
            db,
            feature="chat",
            client=get_client(),
            model=model,
            user_id=user.id if user else None,
            metadata={
                "bucket_id": str(bucket.id),
                "audience": audience,
                "share_id": str(share.id) if share else None,
                "upload_link_id": str(upload_link.id) if upload_link else None,
                "vendor_access_id": str(vendor_access.id) if vendor_access else None,
            },
            max_tokens=1200,
            system=CHAT_SYSTEM,
            messages=[{"role": "user", "content": json.dumps({"context": context, "message": message}, default=str)}],
        )
        parsed = _json_or_fallback(_text_from_response(resp), "answer")
        answer = str(parsed.get("answer") or parsed.get("summary") or _text_from_response(resp))[:5000]
        patch = parsed.get("proposed_context_patch") if isinstance(parsed.get("proposed_context_patch"), dict) else None
        suggested_widget = _suggested_widget_type(parsed)
        assistant = BucketAIMessage(
            bucket_id=bucket.id,
            upload_link_id=upload_link.id if upload_link else None,
            share_id=share.id if share else None,
            vendor_access_id=vendor_access.id if vendor_access else None,
            audience=audience,
            role="assistant",
            author_name="Bucket AI",
            content=answer,
            proposed_context_patch=json_safe_metadata(patch),
            provider="bedrock",
            model=getattr(resp, "model", None) or model,
            metadata_json=json_safe_metadata({"raw": parsed}),
        )
        db.add(assistant)
        await db.flush()
        proposals = await _create_proposals(db, bucket=bucket, source_message=assistant, parsed=parsed, audience=audience, upload_link=upload_link, share=share, vendor_access=vendor_access, user=user)
        await log_bucket_ai_activity(db, bucket.id, "ai_chat_message_created", user=user, actor_name=actor_name, actor_role=audience, target_type="ai_message", target_id=str(user_row.id), detail=message[:180])
        return [user_row, assistant], proposals, suggested_widget
    except Exception as exc:  # noqa: BLE001
        log.exception("bucket_ai: chat failed bucket=%s", bucket.id)
        assistant = BucketAIMessage(
            bucket_id=bucket.id,
            upload_link_id=upload_link.id if upload_link else None,
            share_id=share.id if share else None,
            vendor_access_id=vendor_access.id if vendor_access else None,
            audience=audience,
            role="assistant",
            author_name="Bucket AI",
            content=f"AI is unavailable right now: {exc}",
        )
        db.add(assistant)
        await db.flush()
        await log_bucket_ai_activity(db, bucket.id, "ai_chat_failed", user=user, actor_name=actor_name, actor_role=audience, target_type="ai_message", target_id=str(user_row.id), detail=str(exc)[:180])
        return [user_row, assistant], [], None


async def _chat_context(
    db: AsyncSession,
    *,
    bucket: Bucket,
    audience: str,
    upload_link: BucketUploadLink | None,
    share: BucketShare | None,
    vendor_access: BucketVendorAccess | None,
) -> dict[str, Any]:
    base = {
        "bucket": {
            "name": bucket.name,
            "client_name": bucket.client_name,
            "purpose": bucket.purpose,
            "bucket_type": bucket.bucket_type,
        },
        "audience": audience,
    }
    if audience == "admin":
        review = await latest_review(db, bucket.id)
        tasks = await visible_action_items(db, bucket.id)
        templates = (
            await db.execute(
                select(BucketDocumentTemplate)
                .where(BucketDocumentTemplate.is_active.is_(True))
                .order_by(BucketDocumentTemplate.category, BucketDocumentTemplate.name)
                .limit(250)
            )
        ).scalars().all()
        return {
            **base,
            "ai_context": bucket.ai_context or {},
            "requested_documents": [_doc_context(doc) for doc in bucket.requested_documents],
            "document_template_library": [_template_context(template) for template in templates],
            "files": [_file_context(file) for file in bucket.files if file.status == "uploaded" and file.deleted_at is None],
            "notes": [_note_context(note) for note in bucket.notes],
            "latest_review": review.result if review else None,
            "action_items": [_task_context(task) for task in tasks],
            "instructions": "When the admin asks for tasks or document requests, use the template library where it fits. If no template matches, create a custom action item with route uploader/admin/share.",
        }
    if upload_link is not None:
        review = await latest_review(db, bucket.id)
        tasks = await visible_action_items(db, bucket.id, route="uploader", upload_link_id=upload_link.id, approved_only=True)
        public_ai_context = _public_ai_context(bucket)
        public_review_type = public_ai_context.get("review_type") if public_ai_context else None
        latest_result = review.result if review and isinstance(review.result, dict) else None
        return {
            **base,
            "recipient_name": upload_link.recipient_name,
            "ai_context": public_ai_context or None,
            "requested_documents": [_doc_context(doc) for doc in bucket.requested_documents],
            "uploaded_files": [_file_context(file) for file in bucket.files if file.status == "uploaded" and file.deleted_at is None],
            "visible_summary": upload_link_visible_summary(review, bucket),
            "latest_review": latest_result,
            "document_evidence_map": latest_result.get("document_evidence_map") if latest_result else None,
            "next_best_action": latest_result.get("next_best_action") if latest_result else None,
            "baseline_coverage": (latest_result.get("document_evidence_map") or {}).get("baseline_coverage") if latest_result else None,
            "instructions": (
                "Help the uploader understand what is already uploaded, what is still needed, and how to submit files. "
                "Do not discuss admin notes or shares. External users cannot change saved AI instructions."
                + (
                    " This is a dealer_gatekeeper_v1 file: speak like a human banking underwriter with car dealership finance knowledge, "
                    "keep the baseline package strict, and use the in-chat tool when the user asks to upload files or enter real estate collateral."
                    if public_review_type == "dealer_gatekeeper_v1"
                    else ""
                )
                + (
                    " This is a real_estate_dscr_v1 file: speak like a senior DSCR/investor lending underwriter. Focus on rent support, PITIA, DSCR, LTV, payoff or purchase evidence, property value, entity/vesting, and credit tier. Never ask dealer-specific questions."
                    if public_review_type == "real_estate_dscr_v1"
                    else ""
                )
            ),
            "approved_tasks": [_task_context(task) for task in tasks],
        }
    if share is not None:
        review = await latest_review(db, bucket.id)
        tasks = await visible_action_items(db, bucket.id, route="share", share_id=share.id, approved_only=True)
        notes = [note for note in bucket.notes if note.visibility == "shared" or share.can_see_internal_notes]
        public_ai_context = _public_ai_context(bucket)
        public_review_type = public_ai_context.get("review_type") if public_ai_context else None
        return {
            **base,
            "recipient_name": share.recipient_name,
            "ai_context": public_ai_context or None,
            "visible_files": [_file_context(file) for file in share.files if file.status == "uploaded" and file.deleted_at is None],
            "visible_notes": [_note_context(note) for note in notes],
            "visible_summary": share_visible_summary(review, share) if share.can_view_ai_summary else None,
            "approved_tasks": [_task_context(task) for task in tasks],
            "instructions": (
                "Answer only from the files and notes visible to this share link."
                + (
                    " This is a dealer_gatekeeper_v1 file: answer like a banking underwriter who understands car dealer financing, "
                    "floorplan/MCA exposure, related LLCs, and real estate collateral."
                    if public_review_type == "dealer_gatekeeper_v1"
                    else ""
                )
                + (
                    " This is a real_estate_dscr_v1 file: answer like a DSCR/investor lending underwriter. Focus only on the shared real-estate funding evidence and do not ask dealer-specific questions."
                    if public_review_type == "real_estate_dscr_v1"
                    else ""
                )
            ),
        }
    if vendor_access is not None:
        review = await latest_review(db, bucket.id)
        files = (
            bucket.files if vendor_access.file_scope == "all_active" else vendor_access.files
        )
        visible_files = [
            file for file in files
            if file.status == "uploaded" and file.deleted_at is None
        ]
        tasks = await visible_action_items(db, bucket.id, route="vendor", vendor_access_id=vendor_access.id, approved_only=True)
        notes = [
            note for note in bucket.notes
            if note.visibility == "shared"
            or note.vendor_access_id == vendor_access.id
            or vendor_access.can_see_internal_notes
        ]
        public_ai_context = _public_ai_context(bucket)
        public_review_type = public_ai_context.get("review_type") if public_ai_context else None
        return {
            **base,
            "recipient_name": vendor_access.vendor.name if vendor_access.vendor else "Vendor",
            "ai_context": public_ai_context or None,
            "visible_files": [_file_context(file) for file in visible_files],
            "visible_notes": [_note_context(note) for note in notes],
            "visible_summary": vendor_visible_summary(review, vendor_access) if vendor_access.can_view_ai_summary else None,
            "approved_tasks": [_task_context(task) for task in tasks],
            "instructions": (
                "Answer only from the files and notes visible to this authenticated vendor access. "
                "If the vendor asks for requirements or clarifications, propose tasks for super-admin approval before they become visible to clients or other users."
                + (
                    " This is a dealer_gatekeeper_v1 file: answer like a banking underwriter who understands car dealer financing, "
                    "floorplan/MCA exposure, related LLCs, and real estate collateral."
                    if public_review_type == "dealer_gatekeeper_v1"
                    else ""
                )
                + (
                    " This is a real_estate_dscr_v1 file: answer like a DSCR/investor lending underwriter. Focus only on assigned real-estate funding evidence and do not ask dealer-specific questions."
                    if public_review_type == "real_estate_dscr_v1"
                    else ""
                )
            ),
        }
    return base


async def _create_proposals(
    db: AsyncSession,
    *,
    bucket: Bucket,
    source_message: BucketAIMessage,
    parsed: dict[str, Any],
    audience: str,
    upload_link: BucketUploadLink | None,
    share: BucketShare | None,
    vendor_access: BucketVendorAccess | None,
    user: User | None = None,
) -> list[BucketAIActionItem]:
    raw_items = parsed.get("proposed_action_items")
    if not isinstance(raw_items, list):
        return []
    created: list[BucketAIActionItem] = []
    auto_approve = audience == "admin"
    for raw in raw_items[:5]:
        if not isinstance(raw, dict):
            continue
        title = str(raw.get("title") or "").strip()
        instructions = str(raw.get("instructions") or "").strip()
        if not title or not instructions:
            continue
        route = str(raw.get("route") or ("vendor" if vendor_access else "share" if share else "uploader" if upload_link else "admin")).strip()
        if route not in {"admin", "uploader", "share", "vendor"}:
            route = "admin"
        if route == "share" and share is None and audience == "admin":
            route = "admin"
        if route == "vendor" and vendor_access is None and audience == "admin":
            route = "admin"
        approved_at = _now() if auto_approve else None
        item = BucketAIActionItem(
            bucket_id=bucket.id,
            source_message_id=source_message.id,
            upload_link_id=upload_link.id if route == "uploader" and upload_link else None,
            share_id=share.id if route == "share" and share else None,
            vendor_access_id=vendor_access.id if route == "vendor" and vendor_access else None,
            status="approved" if auto_approve else "proposed",
            route=route,
            title=title[:220],
            instructions=instructions,
            rationale=str(raw.get("rationale") or "")[:2000] or None,
            created_by="ai",
            created_by_user_id=user.id if user else None,
            approved_by_user_id=user.id if approved_at and user else None,
            approved_at=approved_at,
        )
        db.add(item)
        created.append(item)
    if created:
        await db.flush()
        await log_bucket_ai_activity(
            db,
            bucket.id,
            "ai_action_created" if auto_approve else "ai_action_proposed",
            user=user if auto_approve else None,
            actor_name="Bucket AI" if not auto_approve else None,
            actor_role=audience,
            target_type="ai_action_item",
            detail=f"{len(created)} {'created' if auto_approve else 'proposed'}",
        )
    return created


async def _create_review_recommendation_actions(
    db: AsyncSession,
    *,
    bucket: Bucket,
    review: BucketAIReview,
    result: dict[str, Any],
) -> list[BucketAIActionItem]:
    raw_items = result.get("recommended_next_document_requests")
    if not isinstance(raw_items, list):
        return []
    existing_titles = {
        title.lower()
        for title in (
            await db.execute(
                select(BucketAIActionItem.title).where(
                    BucketAIActionItem.bucket_id == bucket.id,
                    BucketAIActionItem.status.in_(("proposed", "approved")),
                )
            )
        ).scalars().all()
    }
    created: list[BucketAIActionItem] = []
    for raw in raw_items[:8]:
        if not isinstance(raw, dict):
            continue
        title = str(raw.get("title") or "").strip()
        instructions = str(raw.get("instructions") or raw.get("detail") or "").strip()
        if not title or not instructions or title.lower() in existing_titles:
            continue
        route = str(raw.get("route") or "admin").strip()
        if route not in {"admin", "uploader", "share"}:
            route = "admin"
        if route == "share":
            route = "admin"
        item = BucketAIActionItem(
            bucket_id=bucket.id,
            source_message_id=None,
            status="proposed",
            route=route,
            title=title[:220],
            instructions=instructions,
            rationale=str(raw.get("rationale") or "Recommended by the latest AI underwriting review.")[:2000],
            created_by="ai",
        )
        db.add(item)
        created.append(item)
        existing_titles.add(title.lower())
    if created:
        await db.flush()
        await log_bucket_ai_activity(
            db,
            bucket.id,
            "ai_action_proposed",
            actor_name="Bucket AI",
            actor_role="system",
            target_type="ai_review",
            target_id=str(review.id),
            detail=f"{len(created)} review recommendations",
        )
    return created


def _doc_context(doc: BucketRequestedDocument) -> dict[str, Any]:
    return {
        "id": str(doc.id),
        "name": doc.name,
        "category": doc.category,
        "description": doc.description,
        "required": doc.required,
        "allow_multiple_files": doc.allow_multiple_files,
        "status": doc.status,
    }


def _file_context(file: BucketFile) -> dict[str, Any]:
    return {
        "id": str(file.id),
        "file_name": file.file_name,
        "content_type": file.content_type,
        "size_bytes": file.size_bytes,
        "requested_document_id": str(file.requested_document_id) if file.requested_document_id else None,
        "uploaded_by_name": file.uploaded_by_name,
    }


def _template_context(template: BucketDocumentTemplate) -> dict[str, Any]:
    return {
        "id": str(template.id),
        "name": template.name,
        "category": template.category,
        "description": template.description,
        "required": template.required,
        "allow_multiple_files": template.allow_multiple_files,
    }


def _note_context(note: BucketNote) -> dict[str, Any]:
    return {"author_name": note.author_name, "visibility": note.visibility, "content": note.content, "created_at": note.created_at.isoformat()}


def _task_context(task: BucketAIActionItem) -> dict[str, Any]:
    return {
        "id": str(task.id),
        "status": task.status,
        "route": task.route,
        "title": task.title,
        "instructions": task.instructions,
        "rationale": task.rationale,
    }
